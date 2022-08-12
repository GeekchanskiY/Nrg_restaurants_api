from django.shortcuts import render, redirect, HttpResponse
from django.http import FileResponse
from .forms import NewUserForm, UploadDataForm
from .models import Dish, DishesCategory, DishSet
from django.contrib.auth import login
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
import openpyxl
import xlsxwriter


def register(request):
    if request.method == "POST":
        form = NewUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registration successful." )
            return HttpResponse("Registered")
        messages.error(request, "Unsuccessful registration. Invalid information.")
    form = NewUserForm()
    return render(request=request, template_name="register.html", context={"register_form": form})


def save_uploaded_file(file):
    with open('data.xlsx', 'wb+') as f:
        for chunk in file.chunks():
            f.write(chunk)


def import_view(request):
    if request.method == "POST":
        form = UploadDataForm(request.POST, request.FILES)
        if form.is_valid():
            # P.S. BytesIO can't be used because of the client's wish of seeing last table using ftp
            save_uploaded_file(request.FILES['table'])
            workbook = openpyxl.load_workbook("data.xlsx")
            sheet = workbook.active
            if sheet.max_row > 1 and sheet.max_column == 7:
                iter_rows = iter(sheet.rows)
                next(iter_rows)
                for row in iter_rows:
                    dish, created = Dish.objects.get_or_create(
                        name_ru=row[0].value,
                        defaults={
                            "name_en": row[1].value,
                            "price": int(row[2].value),
                            "description_ru": row[3].value,
                            "description_en": row[4].value,
                            "restaurant": request.user.restaurant,
                        }
                    )
                    if not created:
                        dish.price = int(row[2].value)
                        dish.save()
                    for category_name in row[5].value.split(";"):
                        try:
                            category_obj = DishesCategory.objects.get(
                                name_ru=category_name,
                                restaurant=request.user.restaurant
                                )
                            if dish not in category_obj.dishes:
                                category_obj.dishes.add(dish)
                        except ObjectDoesNotExist:
                            continue
                    for dish_set_name in row[6].value.split(";"):
                        try:
                            dish_set_obj = DishSet.objects.get(
                                name_ru=dish_set_name,
                                restaurant=request.user.restaurant
                                )
                            if dish not in dish_set_obj.dishes:
                                dish_set_obj.dishes.add(dish)
                        except ObjectDoesNotExist:
                            continue
                return HttpResponse("Updated")
            else:
                return HttpResponse("Таблица создана неправильно (нет строк или не 7 колонок)")
    else:
        form = UploadDataForm()

    return render(request, "upload_file.html", {"upload_file_form": form})


def export_view(request):
    if request.user.restaurant is None:
        return HttpResponse("Вы должны быть авторизованы с аккаунта владельца ресторана")

    dishes = Dish.objects.filter(id=request.user.restaurant.id)

    workbook = xlsxwriter.Workbook("output.xlsx")
    dish_worksheet = workbook.add_worksheet()

    dish_worksheet.write(0, 0, "Название рус")
    dish_worksheet.write(0, 1, "Название англ")
    dish_worksheet.write(0, 2, "цена")
    dish_worksheet.write(0, 3, "Описание рус")
    dish_worksheet.write(0, 4, "Описание англ")
    dish_worksheet.write(0, 5, "Названия категорий на русском через ; (опционально)")
    dish_worksheet.write(0, 6, "Названия наборов на русском через ; (опционально)")

    for row_num, dish in enumerate(dishes):
        dish_worksheet.write(1 + row_num, 0, dish.name_ru)
        dish_worksheet.write(1 + row_num, 1, dish.name_en)
        dish_worksheet.write(1 + row_num, 2, dish.price)
        dish_worksheet.write(1 + row_num, 3, dish.description_ru)
        dish_worksheet.write(1 + row_num, 4, dish.description_en)
        categories = list(category.name_ru for category in dish.dishescategory_set.all().only("name_ru"))
        if categories is not None and len(categories) != 0:
            dish_worksheet.write(1 + row_num, 5, ";".join(categories))
        dish_sets = list(dish_set.name_ru for dish_set in dish.dishset_set.all().only("name_ru"))
        if dish_sets is not None and len(dish_sets) != 0:
            dish_worksheet.write(1 + row_num, 6, ";".join(dish_sets))
    workbook.close()

    return FileResponse(open("output.xlsx", "rb"))
